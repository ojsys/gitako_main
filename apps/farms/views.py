from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Sum, Q, F
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.models import User
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
from .models import (
    Farm, FarmSection, Crop, CropVariety, CropCycle,
    CropCalendar, SeasonalPlanning, PlannedCropAllocation, CropRotationPlan
)
from django.utils import timezone
from datetime import timedelta

# Add these imports at the top
from apps.activities.models import Activity
from apps.financials.models import Budget, BudgetItem

# Inport forms
from apps.farms.forms import (
    FarmSectionForm, FarmSectionFilterForm, CropCycleForm, CropCycleFilterForm, CropForm,
    CropCalendarForm, SeasonalPlanningForm, PlannedCropAllocationForm, CropRotationPlanForm
)


@login_required
def crop_list(request):
    """View to display all crops"""
    # Get all crops
    crops = Crop.objects.all()
    
    # Get user's farms
    farms = Farm.objects.filter(owner=request.user)
    
    # Get active crop cycles for the user's farms
    active_crop_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user,
        status__in=['planned', 'active']
    ).select_related('crop', 'field')
    
    # Count crops by type
    crop_distribution = CropCycle.objects.filter(
        field__farm__owner=request.user,
        status__in=['planned', 'active']
    ).values('crop__name').annotate(
        count=Count('id'),
        total_area=Sum('field__size')
    ).order_by('-total_area')
    
    # Calculate expected yields
    total_expected_yield = CropCycle.objects.filter(
        field__farm__owner=request.user,
        status__in=['planned', 'active']
    ).aggregate(total=Sum('expected_yield_kg'))['total'] or 0
    
    context = {
        'crops': crops,
        'active_crop_cycles': active_crop_cycles,
        'crop_distribution': crop_distribution,
        'total_expected_yield': total_expected_yield,
        'farms': farms,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_list.html', context)

@login_required
def crop_detail(request, crop_id):
    """View to display details of a specific crop"""
    crop = get_object_or_404(Crop, id=crop_id)
    varieties = CropVariety.objects.filter(crop=crop)
    
    # Get user's crop cycles for this crop
    crop_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user,
        crop=crop
    ).select_related('field', 'field__farm')
    
    # Active crop cycles
    active_cycles = crop_cycles.filter(status__in=['planned', 'active'])
    
    # Historical crop cycles
    historical_cycles = crop_cycles.filter(status__in=['harvested', 'failed'])
    
    context = {
        'crop': crop,
        'varieties': varieties,
        'active_cycles': active_cycles,
        'historical_cycles': historical_cycles,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_detail.html', context)


@login_required
def crop_create(request):
    """View to create a new crop."""
    if request.method == 'POST':
        form = CropForm(request.POST, request.FILES)
        if form.is_valid():
            crop = form.save(commit=False)
            # If you have a user associated with the crop, set it here
            # crop.user = request.user # Assuming your Crop model has a user field
            crop.save()
            messages.success(request, f'Crop "{crop.name}" created successfully.')
            return redirect('farms:crop_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CropForm()
    
    context = {
        'form': form,
        'active_page': 'crops',
        'title': 'Add New Crop'
    }
    return render(request, 'farms/crop_form.html', context)

@login_required
def crop_edit(request, crop_id):
    """View to edit an existing crop."""
    crop = get_object_or_404(Crop, id=crop_id)
    # Optional: Add a check here if crops are user-specific
    # if crop.user != request.user:
    #     messages.error(request, "You don't have permission to edit this crop.")
    #     return redirect('farms:crop_list')

    if request.method == 'POST':
        form = CropForm(request.POST, request.FILES, instance=crop)
        if form.is_valid():
            form.save()
            messages.success(request, f'Crop "{crop.name}" updated successfully.')
            return redirect('farms:crop_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CropForm(instance=crop)
    
    context = {
        'form': form,
        'crop': crop,
        'active_page': 'crops',
        'title': f'Edit Crop: {crop.name}'
    }
    return render(request, 'farms/crop_form.html', context)

@login_required
def crop_cycle_list(request):
    """View to display all crop cycles"""
    # Get user's farms
    farms = Farm.objects.filter(owner=request.user)
    
    # Get all crop cycles for the user's farms
    crop_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user
    ).select_related('crop', 'field', 'field__farm')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        crop_cycles = crop_cycles.filter(status=status_filter)
    
    # Filter by farm if provided
    farm_filter = request.GET.get('farm')
    if farm_filter:
        crop_cycles = crop_cycles.filter(field__farm__id=farm_filter)
    
    # Filter by crop if provided
    crop_filter = request.GET.get('crop')
    if crop_filter:
        crop_cycles = crop_cycles.filter(crop__id=crop_filter)
    
    # Get all crops for filtering
    crops = Crop.objects.all()
    
    context = {
        'crop_cycles': crop_cycles,
        'farms': farms,
        'crops': crops,
        'active_page': 'crops',
        'status_filter': status_filter,
        'farm_filter': farm_filter,
        'crop_filter': crop_filter
    }
    
    return render(request, 'farms/crop_cycle_list.html', context)

@login_required
def crop_dashboard(request, crop_cycle_id=None):
    """Dashboard view for crops"""
    if crop_cycle_id:
        crop_cycle = get_object_or_404(CropCycle, id=crop_cycle_id, field__farm__owner=request.user)
        
        # Calculate overall growth stage
        if crop_cycle.planting_date:
            today = timezone.now().date()
            days_since_planting = (today - crop_cycle.planting_date).days
            if crop_cycle.crop.growing_period:
                growth_percentage = min(100, int((days_since_planting / crop_cycle.crop.growing_period) * 100))
            else:
                growth_percentage = 0
        else:
            growth_percentage = 0
        
        # Calculate upcoming harvests (within next 30 days)
        upcoming_harvests = []
        if crop_cycle.status == 'active' and crop_cycle.expected_harvest_date:
            today = timezone.now().date()
            days_until_harvest = (crop_cycle.expected_harvest_date - today).days
            if 0 <= days_until_harvest <= 30:
                upcoming_harvests.append({
                    'crop_name': crop_cycle.crop.name,
                    'field_name': crop_cycle.field.name,
                    'harvest_date': crop_cycle.expected_harvest_date,
                    'days_remaining': days_until_harvest
                })
        

        recent_activities = Activity.objects.filter(crop_cycle=crop_cycle).order_by('-created_at')[:5]
        
        # Get budget data for the dashboard
        # Get active budgets (current date falls between start_date and end_date)
        today = timezone.now().date()
        active_budgets = Budget.objects.filter(
            user=request.user,
            start_date__lte=today,
            end_date__gte=today
        )
        active_budgets_count = active_budgets.count()
        
        # Get recent budgets
        
        recent_budgets = Budget.objects.filter(user=request.user).order_by('-created_at')[:5]
        
        # Calculate total planned income and expenses
        budget_totals = Budget.objects.filter(user=request.user).aggregate(
            total_planned_income=Sum('total_planned_income'),
            total_planned_expenses=Sum('total_planned_expenses')
        )
        
        total_planned_income = budget_totals['total_planned_income'] or 0
        total_planned_expenses = budget_totals['total_planned_expenses'] or 0
        
        context = {
            'crop_cycle': crop_cycle,
            'growth_percentage': growth_percentage,
            'upcoming_harvests': upcoming_harvests,
            'activities': activities,
            'recent_activities': recent_activities,
            # Budget data
            'active_budgets_count': active_budgets_count,
            'recent_budgets': recent_budgets,
            'total_planned_income': total_planned_income,
            'total_planned_expenses': total_planned_expenses,
        }
        
        return render(request, 'farms/crop_dashboard.html', context)
    
    else:
        # Get user's farms
        farms = Farm.objects.filter(owner=request.user)
        
        # Count total crops planted
        total_crops = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status__in=['planned', 'active']
        ).values('crop').distinct().count()
        
        # Get crop distribution
        crop_distribution = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status__in=['planned', 'active']
        ).values('crop__name').annotate(
            total_area=Sum('field__size')
        ).order_by('-total_area')[:4]  # Top 4 crops by area
        
        # Calculate expected yields
        expected_yield = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status__in=['planned', 'active']
        ).aggregate(total=Sum('expected_yield_kg'))['total'] or 0
        
        # Calculate total area under cultivation
        total_area = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status__in=['planned', 'active']
        ).aggregate(total=Sum('field__size'))['total'] or 0
        
        # Determine overall growth stage (simplified)
        now = timezone.now().date()
        active_cycles = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status='active',
            planting_date__lte=now
        )
        
        growth_stage = 'No active crops'
        if active_cycles.exists():
            # Simple logic to determine average growth stage
            early_stage = active_cycles.filter(planting_date__gte=now-timedelta(days=30)).count()
            mid_stage = active_cycles.filter(planting_date__lt=now-timedelta(days=30), 
                                           planting_date__gte=now-timedelta(days=60)).count()
            late_stage = active_cycles.filter(planting_date__lt=now-timedelta(days=60)).count()
            
            if early_stage > mid_stage and early_stage > late_stage:
                growth_stage = 'Early'
            elif mid_stage > early_stage and mid_stage > late_stage:
                growth_stage = 'Mid'
            else:
                growth_stage = 'Late'
        
        # Get upcoming harvests
        upcoming_harvests = CropCycle.objects.filter(
            field__farm__owner=request.user,
            status='active',
            expected_harvest_date__gte=now,
            expected_harvest_date__lte=now+timedelta(days=30)
        ).select_related('crop', 'field')
        
        # Get active budgets (current date falls between start_date and end_date)
        active_budgets = Budget.objects.filter(
            user=request.user,
            # start_date__lte=now,
            # end_date__gte=now
        )
        active_budgets_count = active_budgets.count()
        
        # Get recent activities for this crop cycle
        recent_activities = Activity.objects.filter(created_by=request.user).order_by('-created_at')[:5]
        
        # Get upcoming calendar events (next 7 days)
        upcoming_calendar_events = CropCalendar.objects.filter(
            farm__owner=request.user,
            start_date__gte=now,
            start_date__lte=now + timedelta(days=7),
            status__in=['planned', 'in_progress']
        ).select_related('farm', 'crop_cycle').order_by('start_date', 'priority')[:5]
        
        # Get overdue calendar events
        overdue_calendar_events = CropCalendar.objects.filter(
            farm__owner=request.user,
            start_date__lt=now,
            status__in=['planned', 'in_progress']
        ).select_related('farm', 'crop_cycle').order_by('start_date')[:5]
        
        # Get recent budgets
        recent_budgets = Budget.objects.filter(user=request.user).order_by('-created_at')[:5]
        
        # Calculate total planned income and expenses
        budget_totals = Budget.objects.filter(user=request.user).aggregate(
            total_planned_income=Sum('total_planned_income'),
            total_planned_expenses=Sum('total_planned_expenses')
        )
        
        total_planned_income = budget_totals['total_planned_income'] or 0
        total_planned_expenses = budget_totals['total_planned_expenses'] or 0
        
        context = {
            'farms': farms,
            'total_crops': total_crops,
            'crop_distribution': crop_distribution,
            'expected_yield': expected_yield,
            'growth_stage': growth_stage,
            'upcoming_harvests': upcoming_harvests,
            'active_page': 'crops',
            # Budget data
            'active_budgets_count': active_budgets_count,
            'recent_activities': recent_activities,
            'recent_budgets': recent_budgets,
            'total_planned_income': total_planned_income,
            'total_planned_expenses': total_planned_expenses,
            # Calendar data
            'upcoming_calendar_events': upcoming_calendar_events,
            'overdue_calendar_events': overdue_calendar_events,
        }
        
        return render(request, 'farms/crop_dashboard.html', context)



@login_required
def field_list(request):
    """View to display all fields/farm sections"""
    # Get user's farms
    farms = Farm.objects.filter(owner=request.user)
    
    # Get all fields for the user's farms
    fields = FarmSection.objects.filter(
        farm__owner=request.user,
        is_active=True
    ).select_related('farm').order_by('farm__name', 'name')
    
    # Initialize filter form
    filter_form = FarmSectionFilterForm(request.GET, user=request.user)
    
    # Apply filters if form is valid
    if filter_form.is_valid():
        farm_id = filter_form.cleaned_data.get('farm')
        name = filter_form.cleaned_data.get('name')
        
        if farm_id:
            fields = fields.filter(farm=farm_id)
        
        if name:
            fields = fields.filter(name__icontains=name)
    
    # Calculate totals
    total_fields = fields.count()
    total_area = fields.aggregate(total=Sum('size'))['total'] or 0
    
    # Get active crop cycles count per field
    fields_with_cycles = fields.annotate(
        active_cycles=Count('crop_cycles', filter=Q(crop_cycles__status__in=['planned', 'active']))
    )
    
    context = {
        'fields': fields_with_cycles,
        'farms': farms,
        'total_fields': total_fields,
        'total_area': total_area,
        'filter_form': filter_form,
        'active_page': 'fields'
    }
    
    return render(request, 'farms/field_list.html', context)

@login_required
def field_detail(request, field_id):
    """View to display details of a specific field"""
    field = get_object_or_404(FarmSection, id=field_id, farm__owner=request.user)
    
    # Get crop cycles for this field
    crop_cycles = CropCycle.objects.filter(field=field).select_related('crop', 'crop_variety')
    
    # Active crop cycles
    active_cycles = crop_cycles.filter(status__in=['planned', 'active'])
    
    # Historical crop cycles
    historical_cycles = crop_cycles.filter(status__in=['harvested', 'failed']).order_by('-planting_date')
    
    # Get soil tests for this field
    soil_tests = field.soil_tests.all().order_by('-test_date')[:5]
    
    context = {
        'field': field,
        'active_cycles': active_cycles,
        'historical_cycles': historical_cycles,
        'soil_tests': soil_tests,
        'active_page': 'fields'
    }
    
    return render(request, 'farms/field_detail.html', context)

@login_required
def field_create(request):
    """View to create a new field"""
    # Get user's farms
    farms = Farm.objects.filter(owner=request.user)
    
    if request.method == 'POST':
        form = FarmSectionForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            field = form.save()
            messages.success(request, f'Field "{field.name}" has been created successfully.')
            return redirect('farms:field_detail', field_id=field.id)
    else:
        form = FarmSectionForm(user=request.user)
    
    context = {
        'form': form,
        'farms': farms,
        'active_page': 'fields'
    }
    
    return render(request, 'farms/field_form.html', context)

@login_required
def field_edit(request, field_id):
    """View to edit a field"""
    field = get_object_or_404(FarmSection, id=field_id, farm__owner=request.user)
    
    if request.method == 'POST':
        form = FarmSectionForm(request.POST, request.FILES, instance=field, user=request.user)
        if form.is_valid():
            field = form.save()
            messages.success(request, f'Field "{field.name}" has been updated successfully.')
            return redirect('farms:field_detail', field_id=field.id)
    else:
        form = FarmSectionForm(instance=field, user=request.user)
    
    context = {
        'form': form,
        'field': field,
        'active_page': 'fields'
    }
    
    return render(request, 'farms/field_form.html', context)

@login_required
def field_delete(request, field_id):
    """View to delete a field"""
    field = get_object_or_404(FarmSection, id=field_id, farm__owner=request.user)
    
    if request.method == 'POST':
        field_name = field.name
        field.delete()
        messages.success(request, f'Field "{field_name}" has been deleted successfully.')
        return redirect('farms:field_list')
    
    context = {
        'field': field,
        'active_page': 'fields'
    }
    
    return render(request, 'farms/field_confirm_delete.html', context)

@login_required
def crop_cycle_create(request):
    """View to create a new crop cycle"""
    if request.method == 'POST':
        form = CropCycleForm(request.POST, user=request.user)
        if form.is_valid():
            crop_cycle = form.save(commit=False)
            crop_cycle.farm = request.user.profile.farm  # Assuming farm is linked to user profile
            crop_cycle.save()
            messages.success(request, 'Crop cycle created successfully.')
            return redirect('farms:crop_cycle_detail', pk=crop_cycle.id)  # Corrected line
    else:
        form = CropCycleForm(user=request.user)
        
        # Pre-select field if provided in URL
        field_id = request.GET.get('field')
        if field_id:
            try:
                field = FarmSection.objects.get(id=field_id, farm__owner=request.user)
                form.fields['field'].initial = field
            except FarmSection.DoesNotExist:
                pass
    
    context = {
        'form': form,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_cycle_form.html', context)

@login_required
def crop_cycle_detail(request, cycle_id):
    """View to display details of a specific crop cycle"""
    crop_cycle = get_object_or_404(CropCycle, id=cycle_id, field__farm__owner=request.user)
    
    # Calculate growth progress
    if crop_cycle.planting_date and crop_cycle.expected_harvest_date:
        from django.utils import timezone
        today = timezone.now().date()
        total_days = (crop_cycle.expected_harvest_date - crop_cycle.planting_date).days
        days_elapsed = (today - crop_cycle.planting_date).days
        
        if total_days > 0:
            growth_percentage = min(100, max(0, int((days_elapsed / total_days) * 100)))
        else:
            growth_percentage = 0
    else:
        growth_percentage = 0
    
    # Get related activities (you'll need to import Activity model)
    try:
        from apps.activities.models import Activity
        activities = Activity.objects.filter(crop_cycle=crop_cycle).order_by('-actual_date', '-planned_date')
    except ImportError:
        activities = None
    
    context = {
        'crop_cycle': crop_cycle,
        'growth_percentage': growth_percentage,
        'activities': activities,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_cycle_detail.html', context)

@login_required
def crop_cycle_edit(request, cycle_id):
    """View to edit a crop cycle"""
    crop_cycle = get_object_or_404(CropCycle, id=cycle_id, field__farm__owner=request.user)
    
    if request.method == 'POST':
        form = CropCycleForm(request.POST, instance=crop_cycle, user=request.user)
        if form.is_valid():
            crop_cycle = form.save()
            messages.success(request, f'Crop cycle for {crop_cycle.crop.name} has been updated successfully.')
            return redirect('farms:crop_cycle_detail', cycle_id=crop_cycle.id)
    else:
        form = CropCycleForm(instance=crop_cycle, user=request.user)
    
    context = {
        'form': form,
        'crop_cycle': crop_cycle,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_cycle_form.html', context)

@login_required
def crop_cycle_delete(request, cycle_id):
    """View to delete a crop cycle"""
    crop_cycle = get_object_or_404(CropCycle, id=cycle_id, field__farm__owner=request.user)
    
    if request.method == 'POST':
        crop_name = f"{crop_cycle.crop.name} at {crop_cycle.field.name}"
        crop_cycle.delete()
        messages.success(request, f'Crop cycle "{crop_name}" has been deleted successfully.')
        return redirect('farms:crop_cycle_list')
    
    context = {
        'crop_cycle': crop_cycle,
        'active_page': 'crops'
    }
    
    return render(request, 'farms/crop_cycle_confirm_delete.html', context)

# AJAX views for dynamic form loading
@login_required
def get_crop_varieties(request):
    """AJAX view to get crop varieties based on selected crop"""
    crop_id = request.GET.get('crop_id')
    varieties = []
    
    if crop_id:
        try:
            crop = Crop.objects.get(id=crop_id)
            varieties = list(crop.varieties.values('id', 'name'))
        except Crop.DoesNotExist:
            pass
    
    return JsonResponse({'varieties': varieties})

@login_required
def get_farm_fields(request):
    """AJAX view to get fields based on selected farm"""
    farm_id = request.GET.get('farm_id')
    fields = []
    
    if farm_id:
        try:
            farm = Farm.objects.get(id=farm_id, owner=request.user)
            fields = list(farm.sections.filter(is_active=True).values('id', 'name', 'size'))
        except Farm.DoesNotExist:
            pass
    
    return JsonResponse({'fields': fields})

# Crop Calendar Views
@login_required
def crop_calendar_view(request):
    """View to display the crop calendar"""
    # Get current month/year or from query params
    from django.utils import timezone
    import calendar
    from datetime import datetime, date
    
    today = timezone.now().date()
    current_month = int(request.GET.get('month', today.month))
    current_year = int(request.GET.get('year', today.year))
    
    # Get calendar events for the current month
    start_date = date(current_year, current_month, 1)
    # Get last day of month
    last_day = calendar.monthrange(current_year, current_month)[1]
    end_date = date(current_year, current_month, last_day)
    
    # Get user's calendar events
    calendar_events = CropCalendar.objects.filter(
        farm__owner=request.user,
        start_date__lte=end_date,
        end_date__gte=start_date
    ).select_related('farm', 'crop_cycle', 'assigned_to').order_by('start_date', 'priority')
    
    # Get upcoming events (next 30 days)
    upcoming_events = CropCalendar.objects.filter(
        farm__owner=request.user,
        start_date__gte=today,
        start_date__lte=today + timedelta(days=30)
    ).select_related('farm', 'crop_cycle').order_by('start_date')[:10]
    
    # Get overdue events
    overdue_events = CropCalendar.objects.filter(
        farm__owner=request.user,
        start_date__lt=today,
        status__in=['planned', 'in_progress']
    ).select_related('farm', 'crop_cycle').order_by('start_date')[:10]
    
    # Calendar navigation
    prev_month = current_month - 1 if current_month > 1 else 12
    prev_year = current_year if current_month > 1 else current_year - 1
    next_month = current_month + 1 if current_month < 12 else 1
    next_year = current_year if current_month < 12 else current_year + 1
    
    # Generate calendar data
    cal = calendar.monthcalendar(current_year, current_month)
    month_name = calendar.month_name[current_month]
    
    # Group events by date for calendar display
    events_by_date = {}
    for event in calendar_events:
        event_date = event.start_date
        if event_date not in events_by_date:
            events_by_date[event_date] = []
        events_by_date[event_date].append(event)
    
    context = {
        'calendar_events': calendar_events,
        'upcoming_events': upcoming_events,
        'overdue_events': overdue_events,
        'events_by_date': events_by_date,
        'calendar_data': cal,
        'current_month': current_month,
        'current_year': current_year,
        'month_name': month_name,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'today': today,
        'active_page': 'calendar'
    }
    
    return render(request, 'farms/crop_calendar.html', context)

@login_required
def crop_calendar_create(request):
    """View to create a new calendar event"""
    if request.method == 'POST':
        form = CropCalendarForm(request.POST, user=request.user)
        if form.is_valid():
            calendar_event = form.save(commit=False)
            calendar_event.created_by = request.user
            calendar_event.save()
            messages.success(request, f'Calendar event "{calendar_event.title}" created successfully.')
            return redirect('farms:crop_calendar')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CropCalendarForm(user=request.user)
    
    context = {
        'form': form,
        'active_page': 'calendar',
        'title': 'Add Calendar Event'
    }
    return render(request, 'farms/calendar_event_form.html', context)

@login_required
def crop_calendar_edit(request, event_id):
    """View to edit a calendar event"""
    calendar_event = get_object_or_404(CropCalendar, id=event_id, farm__owner=request.user)
    
    if request.method == 'POST':
        form = CropCalendarForm(request.POST, instance=calendar_event, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f'Calendar event "{calendar_event.title}" updated successfully.')
            return redirect('farms:crop_calendar')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CropCalendarForm(instance=calendar_event, user=request.user)
    
    context = {
        'form': form,
        'calendar_event': calendar_event,
        'active_page': 'calendar',
        'title': f'Edit Event: {calendar_event.title}'
    }
    return render(request, 'farms/calendar_event_form.html', context)

@login_required
def crop_calendar_delete(request, event_id):
    """View to delete a calendar event"""
    calendar_event = get_object_or_404(CropCalendar, id=event_id, farm__owner=request.user)
    
    if request.method == 'POST':
        title = calendar_event.title
        calendar_event.delete()
        messages.success(request, f'Calendar event "{title}" has been deleted successfully.')
        return redirect('farms:crop_calendar')
    
    context = {
        'calendar_event': calendar_event,
        'active_page': 'calendar'
    }
    
    return render(request, 'farms/calendar_event_confirm_delete.html', context)

@login_required
def crop_calendar_complete(request, event_id):
    """AJAX view to mark a calendar event as completed"""
    if request.method == 'POST':
        calendar_event = get_object_or_404(CropCalendar, id=event_id, farm__owner=request.user)
        
        calendar_event.status = 'completed'
        calendar_event.completion_date = timezone.now().date()
        calendar_event.completion_notes = request.POST.get('notes', '')
        calendar_event.actual_cost = request.POST.get('actual_cost') or calendar_event.estimated_cost
        calendar_event.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Event "{calendar_event.title}" marked as completed.'
        })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

# Seasonal Planning Views
@login_required
def seasonal_planning_list(request):
    """View to display seasonal planning"""
    seasonal_plans = SeasonalPlanning.objects.filter(
        farm__owner=request.user
    ).select_related('farm', 'field').order_by('-season_year', 'season')
    
    context = {
        'seasonal_plans': seasonal_plans,
        'active_page': 'planning'
    }
    return render(request, 'farms/seasonal_planning_list.html', context)

@login_required
def seasonal_planning_create(request):
    """View to create a new seasonal plan"""
    if request.method == 'POST':
        form = SeasonalPlanningForm(request.POST, user=request.user)
        if form.is_valid():
            seasonal_plan = form.save(commit=False)
            seasonal_plan.created_by = request.user
            seasonal_plan.save()
            messages.success(request, f'Seasonal plan for {seasonal_plan.field.name} created successfully.')
            return redirect('farms:seasonal_planning_detail', plan_id=seasonal_plan.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SeasonalPlanningForm(user=request.user)
    
    context = {
        'form': form,
        'active_page': 'planning',
        'title': 'Create Seasonal Plan'
    }
    return render(request, 'farms/seasonal_planning_form.html', context)

@login_required
def seasonal_planning_detail(request, plan_id):
    """View to display seasonal plan details"""
    seasonal_plan = get_object_or_404(SeasonalPlanning, id=plan_id, farm__owner=request.user)
    
    # Get planned crop allocations
    crop_allocations = PlannedCropAllocation.objects.filter(
        seasonal_plan=seasonal_plan
    ).select_related('crop', 'crop_variety')
    
    context = {
        'seasonal_plan': seasonal_plan,
        'crop_allocations': crop_allocations,
        'active_page': 'planning'
    }
    return render(request, 'farms/seasonal_planning_detail.html', context)

# Crop Rotation Views
@login_required
def crop_rotation_list(request):
    """View to display crop rotation plans"""
    rotation_plans = CropRotationPlan.objects.filter(
        farm__owner=request.user
    ).select_related('farm', 'field', 'year_1_crop', 'year_2_crop', 'year_3_crop', 'year_4_crop')
    
    context = {
        'rotation_plans': rotation_plans,
        'active_page': 'rotation'
    }
    return render(request, 'farms/crop_rotation_list.html', context)

@login_required
def crop_rotation_create(request):
    """View to create a new crop rotation plan"""
    if request.method == 'POST':
        form = CropRotationPlanForm(request.POST, user=request.user)
        if form.is_valid():
            rotation_plan = form.save(commit=False)
            rotation_plan.created_by = request.user
            rotation_plan.save()
            messages.success(request, f'Crop rotation plan "{rotation_plan.plan_name}" created successfully.')
            return redirect('farms:crop_rotation_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CropRotationPlanForm(user=request.user)
    
    context = {
        'form': form,
        'active_page': 'rotation',
        'title': 'Create Crop Rotation Plan'
    }
    return render(request, 'farms/crop_rotation_form.html', context)

@login_required
def crop_rotation_recommendations(request):
    """View to show AI-powered crop rotation recommendations"""
    user_farms = Farm.objects.filter(owner=request.user, is_active=True)
    
    # Get fields without active rotation plans
    fields_without_rotation = FarmSection.objects.filter(
        farm__owner=request.user,
        is_active=True
    ).exclude(
        rotation_plans__is_active=True
    ).select_related('farm')
    
    # Get recent crop history for each field
    field_history = {}
    for field in fields_without_rotation:
        recent_cycles = CropCycle.objects.filter(
            field=field,
            status='harvested'
        ).select_related('crop').order_by('-actual_harvest_date')[:3]
        field_history[field.id] = recent_cycles
    
    # Simple rotation recommendations (can be enhanced with AI later)
    recommendations = []
    for field in fields_without_rotation:
        recent_crops = field_history.get(field.id, [])
        if recent_crops:
            last_crop = recent_crops[0].crop
            # Simple logic: recommend different crop families
            recommended_crops = Crop.objects.exclude(id=last_crop.id)[:3]
        else:
            # Default recommendations for new fields
            recommended_crops = Crop.objects.all()[:3]
        
        recommendations.append({
            'field': field,
            'recent_crops': recent_crops,
            'recommended_crops': recommended_crops,
            'reasoning': generate_rotation_reasoning(recent_crops, recommended_crops)
        })
    
    context = {
        'recommendations': recommendations,
        'fields_without_rotation': fields_without_rotation,
        'active_page': 'rotation'
    }
    return render(request, 'farms/crop_rotation_recommendations.html', context)

def generate_rotation_reasoning(recent_crops, recommended_crops):
    """Generate simple reasoning for crop rotation recommendations"""
    if not recent_crops:
        return "This field is new, so we recommend starting with nitrogen-fixing legumes to improve soil health."
    
    last_crop = recent_crops[0].crop
    reasons = []
    
    # Simple logic - can be enhanced
    if 'nitrogen' in last_crop.name.lower() or 'bean' in last_crop.name.lower() or 'pea' in last_crop.name.lower():
        reasons.append("Previous legume crop has enriched soil with nitrogen")
        reasons.append("Good time to plant heavy feeders like corn or leafy greens")
    elif 'corn' in last_crop.name.lower() or 'maize' in last_crop.name.lower():
        reasons.append("Corn is a heavy feeder, soil needs restoration")
        reasons.append("Legumes will replenish nitrogen levels")
    else:
        reasons.append("Rotation will break pest and disease cycles")
        reasons.append("Different root depths will improve soil structure")
    
    return "; ".join(reasons)

# Yield Analytics Views
@login_required
def yield_analytics(request):
    """View to display predictive yield analytics"""
    from django.db.models import Avg, Sum, Count, Max, Min
    from datetime import datetime, timedelta
    import json
    
    user_farms = Farm.objects.filter(owner=request.user, is_active=True)
    
    # Get historical crop cycles with actual yields
    historical_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user,
        status='harvested',
        actual_yield_kg__isnull=False
    ).select_related('crop', 'field', 'field__farm').order_by('-actual_harvest_date')
    
    # Yield trends by crop
    crop_yield_trends = historical_cycles.values(
        'crop__name'
    ).annotate(
        avg_yield=Avg('actual_yield_kg'),
        total_cycles=Count('id'),
        max_yield=Max('actual_yield_kg'),
        min_yield=Min('actual_yield_kg'),
        total_area=Sum('field__size')
    ).order_by('-avg_yield')
    
    # Monthly yield patterns
    monthly_yields = {}
    for cycle in historical_cycles:
        if cycle.actual_harvest_date:
            month = cycle.actual_harvest_date.strftime('%B')
            crop = cycle.crop.name
            if month not in monthly_yields:
                monthly_yields[month] = {}
            if crop not in monthly_yields[month]:
                monthly_yields[month][crop] = []
            monthly_yields[month][crop].append(float(cycle.actual_yield_kg))
    
    # Field performance comparison
    field_performance = historical_cycles.values(
        'field__name', 'field__farm__name'
    ).annotate(
        avg_yield=Avg('actual_yield_kg'),
        total_harvest=Sum('actual_yield_kg'),
        cycle_count=Count('id')
    ).order_by('-avg_yield')[:10]
    
    # Yield predictions for active cycles
    active_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user,
        status='active',
        expected_harvest_date__isnull=False
    ).select_related('crop', 'field')
    
    yield_predictions = []
    for cycle in active_cycles:
        # Simple prediction based on historical averages
        historical_avg = historical_cycles.filter(
            crop=cycle.crop
        ).aggregate(avg_yield=Avg('actual_yield_kg'))['avg_yield'] or 0
        
        # Factor in field size for total prediction
        predicted_total = float(historical_avg) * float(cycle.field.size)
        
        yield_predictions.append({
            'cycle': cycle,
            'predicted_yield_kg': predicted_total,
            'historical_avg': historical_avg,
            'confidence': calculate_prediction_confidence(cycle, historical_cycles)
        })
    
    # Performance insights
    insights = generate_yield_insights(historical_cycles, crop_yield_trends)
    
    context = {
        'historical_cycles': historical_cycles[:20],  # Last 20 harvests
        'crop_yield_trends': crop_yield_trends,
        'monthly_yields': json.dumps(monthly_yields),
        'field_performance': field_performance,
        'yield_predictions': yield_predictions,
        'insights': insights,
        'total_harvested': historical_cycles.aggregate(Sum('actual_yield_kg'))['actual_yield_kg__sum'] or 0,
        'avg_yield_per_hectare': calculate_avg_yield_per_hectare(historical_cycles),
        'active_page': 'analytics'
    }
    
    return render(request, 'farms/yield_analytics.html', context)

def calculate_prediction_confidence(cycle, historical_cycles):
    """Calculate confidence level for yield prediction"""
    similar_cycles = historical_cycles.filter(crop=cycle.crop).count()
    
    if similar_cycles >= 10:
        return 'High'
    elif similar_cycles >= 5:
        return 'Medium'
    elif similar_cycles >= 2:
        return 'Low'
    else:
        return 'Very Low'

def calculate_avg_yield_per_hectare(historical_cycles):
    """Calculate average yield per hectare across all crops"""
    total_yield = 0
    total_area = 0
    
    for cycle in historical_cycles:
        if cycle.actual_yield_kg and cycle.field.size:
            total_yield += float(cycle.actual_yield_kg)
            total_area += float(cycle.field.size)
    
    return round(total_yield / total_area, 2) if total_area > 0 else 0

def generate_yield_insights(historical_cycles, crop_yield_trends):
    """Generate insights from yield data"""
    insights = []
    
    if not historical_cycles:
        insights.append({
            'type': 'info',
            'title': 'Start Tracking Yields',
            'message': 'Begin recording actual harvest yields to generate predictive analytics.'
        })
        return insights
    
    # Best performing crop
    if crop_yield_trends:
        best_crop = crop_yield_trends[0]
        insights.append({
            'type': 'success',
            'title': 'Top Performer',
            'message': f"{best_crop['crop__name']} shows the highest average yield at {best_crop['avg_yield']:.1f} kg."
        })
    
    # Seasonal patterns
    current_month = datetime.now().strftime('%B')
    recent_yields = historical_cycles.filter(
        actual_harvest_date__month=datetime.now().month
    )
    
    if recent_yields:
        avg_current_month = recent_yields.aggregate(Avg('actual_yield_kg'))['actual_yield_kg__avg']
        insights.append({
            'type': 'info',
            'title': f'{current_month} Performance',
            'message': f"Historical {current_month} harvests average {avg_current_month:.1f} kg."
        })
    
    # Improvement opportunity
    low_yield_crops = [trend for trend in crop_yield_trends if trend['total_cycles'] >= 3]
    if len(low_yield_crops) > 1:
        lowest = low_yield_crops[-1]
        highest = low_yield_crops[0]
        improvement = ((highest['avg_yield'] - lowest['avg_yield']) / lowest['avg_yield']) * 100
        
        insights.append({
            'type': 'warning',
            'title': 'Improvement Opportunity',
            'message': f"{lowest['crop__name']} could potentially improve by {improvement:.0f}% based on {highest['crop__name']} performance."
        })
    
    return insights

@login_required
def crop_performance_detail(request, crop_id):
    """Detailed performance view for a specific crop"""
    from django.db.models import Avg, Count, Sum
    import json
    
    crop = get_object_or_404(Crop, id=crop_id)
    
    # Get all cycles for this crop
    crop_cycles = CropCycle.objects.filter(
        field__farm__owner=request.user,
        crop=crop,
        status='harvested',
        actual_yield_kg__isnull=False
    ).select_related('field', 'field__farm').order_by('-actual_harvest_date')
    
    if not crop_cycles:
        messages.info(request, f'No harvest data available for {crop.name} yet.')
        return redirect('farms:yield_analytics')
    
    # Performance statistics
    stats = crop_cycles.aggregate(
        avg_yield=Avg('actual_yield_kg'),
        total_yield=Sum('actual_yield_kg'),
        total_cycles=Count('id'),
        max_yield=Max('actual_yield_kg'),
        min_yield=Min('actual_yield_kg')
    )
    
    # Yield by field
    field_performance = crop_cycles.values(
        'field__name', 'field__farm__name'
    ).annotate(
        avg_yield=Avg('actual_yield_kg'),
        cycle_count=Count('id')
    ).order_by('-avg_yield')
    
    # Yield over time (for chart)
    yield_timeline = []
    for cycle in crop_cycles:
        yield_timeline.append({
            'date': cycle.actual_harvest_date.strftime('%Y-%m-%d'),
            'yield': float(cycle.actual_yield_kg),
            'field': cycle.field.name
        })
    
    # Growing season analysis
    season_performance = {}
    for cycle in crop_cycles:
        season = get_season_from_date(cycle.planting_date)
        if season not in season_performance:
            season_performance[season] = {'yields': [], 'count': 0}
        season_performance[season]['yields'].append(float(cycle.actual_yield_kg))
        season_performance[season]['count'] += 1
    
    # Calculate averages for seasons
    for season in season_performance:
        yields = season_performance[season]['yields']
        season_performance[season]['avg'] = sum(yields) / len(yields)
    
    context = {
        'crop': crop,
        'crop_cycles': crop_cycles,
        'stats': stats,
        'field_performance': field_performance,
        'yield_timeline': json.dumps(yield_timeline),
        'season_performance': season_performance,
        'active_page': 'analytics'
    }
    
    return render(request, 'farms/crop_performance_detail.html', context)

def get_season_from_date(date):
    """Determine season from planting date"""
    month = date.month
    if month in [12, 1, 2]:
        return 'Winter'
    elif month in [3, 4, 5]:
        return 'Spring'
    elif month in [6, 7, 8]:
        return 'Summer'
    else:
        return 'Fall'

# Excel Template and Bulk Upload Views

@login_required
def crop_calendar_download_template(request):
    """Generate and download Excel template for bulk calendar events upload"""
    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Calendar Events Template"
    
    # Define headers
    headers = [
        'Farm Name', 'Field Name', 'Event Type', 'Title', 'Description',
        'Start Date', 'End Date', 'Priority', 'Status', 'Is Recurring',
        'Recurrence Pattern', 'Recurrence Interval', 'Assigned To Email',
        'Weather Dependent', 'Min Temperature', 'Max Temperature', 
        'Max Wind Speed', 'No Rain Required', 'Estimated Cost'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 15
    
    # Add data validation sheet
    validation_sheet = workbook.create_sheet("Reference Data")
    
    # Get user's farms and fields
    user_farms = Farm.objects.filter(owner=request.user)
    farm_names = [farm.name for farm in user_farms]
    field_names = []
    for farm in user_farms:
        for field in farm.sections.all():
            field_names.append(f"{farm.name} - {field.name}")
    
    # Event types
    event_types = ['planting', 'fertilizing', 'weeding', 'watering', 'pest_control', 
                   'pruning', 'harvesting', 'soil_preparation', 'maintenance', 'other']
    
    # Priority levels
    priorities = ['low', 'medium', 'high', 'critical']
    
    # Status options
    statuses = ['planned', 'in_progress', 'completed', 'cancelled', 'overdue']
    
    # Recurrence patterns
    recurrence_patterns = ['daily', 'weekly', 'monthly']
    
    # Add reference data to validation sheet
    validation_data = {
        'Farm Names': farm_names,
        'Field Names': field_names,
        'Event Types': event_types,
        'Priorities': priorities,
        'Statuses': statuses,
        'Recurrence Patterns': recurrence_patterns
    }
    
    col = 1
    for category, items in validation_data.items():
        validation_sheet.cell(row=1, column=col, value=category).font = Font(bold=True)
        for row, item in enumerate(items, 2):
            validation_sheet.cell(row=row, column=col, value=item)
        col += 1
    
    # Add example rows to main sheet
    example_data = [
        [
            'Main Farm', 'North Field', 'planting', 'Plant Maize', 
            'Plant maize seeds in north field', '2024-03-15', '2024-03-17',
            'high', 'planned', 'FALSE', '', '1', 'farmer@example.com',
            'TRUE', '15', '35', '20', 'TRUE', '500.00'
        ],
        [
            'Main Farm', 'South Field', 'fertilizing', 'Apply NPK Fertilizer',
            'Apply NPK fertilizer to boost crop growth', '2024-04-01', '',
            'medium', 'planned', 'FALSE', '', '1', '',
            'FALSE', '', '', '', 'FALSE', '200.00'
        ]
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.border = border
    
    # Add instructions sheet
    instructions_sheet = workbook.create_sheet("Instructions")
    instructions = [
        "CROP CALENDAR BULK UPLOAD INSTRUCTIONS",
        "",
        "1. REQUIRED FIELDS (must be filled):",
        "   - Farm Name: Must match exactly with your existing farm names",
        "   - Field Name: Must match exactly with your existing field names", 
        "   - Event Type: Choose from the dropdown list",
        "   - Title: Brief description of the event",
        "   - Start Date: Format as YYYY-MM-DD (e.g., 2024-03-15)",
        "",
        "2. OPTIONAL FIELDS:",
        "   - Description: Detailed event description",
        "   - End Date: For multi-day events (format YYYY-MM-DD)",
        "   - Priority: low, medium, high, critical (default: medium)",
        "   - Status: planned, in_progress, completed, cancelled (default: planned)",
        "   - Is Recurring: TRUE/FALSE (default: FALSE)",
        "   - Recurrence Pattern: daily, weekly, monthly (if recurring)",
        "   - Recurrence Interval: Number (e.g., 1 for every week, 2 for every 2 weeks)",
        "   - Assigned To Email: Email of person assigned to the task",
        "   - Weather Dependent: TRUE/FALSE",
        "   - Min/Max Temperature: Minimum and maximum temperature in Celsius",
        "   - Max Wind Speed: Maximum wind speed in km/h",
        "   - No Rain Required: TRUE/FALSE",
        "   - Estimated Cost: Cost in currency units",
        "",
        "3. IMPORTANT NOTES:",
        "   - Delete the example rows before uploading",
        "   - Farm and Field names must exist in your account",
        "   - Dates must be in YYYY-MM-DD format",
        "   - Boolean fields use TRUE/FALSE",
        "   - Email addresses must be valid registered users",
        "   - Save the file as Excel (.xlsx) format",
        "",
        "4. UPLOAD PROCESS:",
        "   - Go to Crop Calendar page",
        "   - Click 'Bulk Upload' button",
        "   - Select your completed Excel file",
        "   - Review the upload summary",
        "   - Confirm to create events",
        "",
        "5. VALIDATION:",
        "   - Invalid data will be highlighted in red",
        "   - You can fix errors and re-upload",
        "   - Only valid rows will be processed"
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row, column=1, value=instruction)
        if row == 1:  # Title
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith(('1.', '2.', '3.', '4.', '5.')):  # Section headers
            cell.font = Font(bold=True, size=12)
    
    instructions_sheet.column_dimensions['A'].width = 80
    
    # Prepare response
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="crop_calendar_template.xlsx"'
    
    return response

@login_required
def crop_calendar_bulk_upload(request):
    """Handle bulk upload of calendar events from Excel file"""
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file uploaded. Please select an Excel file.')
            return redirect('farms:crop_calendar')
        
        file = request.FILES['excel_file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
            return redirect('farms:crop_calendar')
        
        try:
            # Read the Excel file
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
            
            # Skip header row and example rows
            data_rows = list(worksheet.iter_rows(min_row=4, values_only=True))
            
            # Filter out empty rows
            data_rows = [row for row in data_rows if any(cell for cell in row)]
            
            if not data_rows:
                messages.warning(request, 'No data found in the Excel file. Please add events to upload.')
                return redirect('farms:crop_calendar')
            
            # Validate and process data
            valid_events = []
            errors = []
            
            for row_num, row in enumerate(data_rows, 4):  # Start from row 4 (after headers and examples)
                try:
                    event_data = process_calendar_row(row, request.user, row_num)
                    if event_data:
                        valid_events.append(event_data)
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            # Create events if validation passes
            if valid_events and not errors:
                created_count = 0
                for event_data in valid_events:
                    CropCalendar.objects.create(**event_data)
                    created_count += 1
                
                messages.success(request, f'Successfully created {created_count} calendar events.')
                
            elif valid_events and errors:
                # Partial success
                created_count = 0
                for event_data in valid_events:
                    CropCalendar.objects.create(**event_data)
                    created_count += 1
                
                messages.warning(
                    request, 
                    f'Created {created_count} events successfully. {len(errors)} rows had errors: {"; ".join(errors[:3])}'
                )
            else:
                messages.error(request, f'Upload failed. Errors found: {"; ".join(errors[:5])}')
                
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
    
    return redirect('farms:crop_calendar')

def process_calendar_row(row, user, row_num):
    """Process a single row from Excel file and return validated event data"""
    if not row or len(row) < 6:  # Minimum required fields
        return None
    
    # Map row data to fields
    farm_name = row[0]
    field_name = row[1] 
    event_type = row[2]
    title = row[3]
    description = row[4] or ''
    start_date = row[5]
    end_date = row[6] if len(row) > 6 and row[6] else None
    priority = row[7] if len(row) > 7 and row[7] else 'medium'
    status = row[8] if len(row) > 8 and row[8] else 'planned'
    is_recurring = row[9] if len(row) > 9 else False
    recurrence_pattern = row[10] if len(row) > 10 and row[10] else None
    recurrence_interval = row[11] if len(row) > 11 and row[11] else 1
    assigned_to_email = row[12] if len(row) > 12 and row[12] else None
    weather_dependent = row[13] if len(row) > 13 else False
    min_temperature = row[14] if len(row) > 14 and row[14] else None
    max_temperature = row[15] if len(row) > 15 and row[15] else None
    max_wind_speed = row[16] if len(row) > 16 and row[16] else None
    no_rain_required = row[17] if len(row) > 17 else False
    estimated_cost = row[18] if len(row) > 18 and row[18] else None
    
    # Validate required fields
    if not all([farm_name, event_type, title, start_date]):
        raise ValueError("Missing required fields: Farm Name, Event Type, Title, or Start Date")
    
    # Find farm
    try:
        farm = Farm.objects.get(name=farm_name, owner=user)
    except Farm.DoesNotExist:
        raise ValueError(f"Farm '{farm_name}' not found")
    
    # Parse date
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("Start date must be in YYYY-MM-DD format")
    
    if end_date and isinstance(end_date, str):
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("End date must be in YYYY-MM-DD format")
    
    # Validate choices
    valid_event_types = ['planting', 'fertilizing', 'weeding', 'watering', 'pest_control', 
                        'pruning', 'harvesting', 'soil_preparation', 'maintenance', 'other']
    if event_type not in valid_event_types:
        raise ValueError(f"Invalid event type: {event_type}")
    
    valid_priorities = ['low', 'medium', 'high', 'critical']
    if priority not in valid_priorities:
        raise ValueError(f"Invalid priority: {priority}")
    
    valid_statuses = ['planned', 'in_progress', 'completed', 'cancelled', 'overdue']
    if status not in valid_statuses:
        raise ValueError(f"Invalid status: {status}")
    
    # Find assigned user if email provided
    assigned_to = None
    if assigned_to_email:
        try:
            assigned_to = User.objects.get(email=assigned_to_email)
        except User.DoesNotExist:
            raise ValueError(f"User with email '{assigned_to_email}' not found")
    
    # Convert boolean fields
    if isinstance(is_recurring, str):
        is_recurring = is_recurring.upper() == 'TRUE'
    if isinstance(weather_dependent, str):
        weather_dependent = weather_dependent.upper() == 'TRUE'  
    if isinstance(no_rain_required, str):
        no_rain_required = no_rain_required.upper() == 'TRUE'
    
    # Build event data
    event_data = {
        'farm': farm,
        'event_type': event_type,
        'title': title,
        'description': description,
        'start_date': start_date,
        'end_date': end_date,
        'priority': priority,
        'status': status,
        'is_recurring': bool(is_recurring),
        'recurrence_pattern': recurrence_pattern,
        'recurrence_interval': int(recurrence_interval) if recurrence_interval else 1,
        'assigned_to': assigned_to,
        'weather_dependent': bool(weather_dependent),
        'min_temperature': float(min_temperature) if min_temperature else None,
        'max_temperature': float(max_temperature) if max_temperature else None,
        'max_wind_speed': float(max_wind_speed) if max_wind_speed else None,
        'no_rain_required': bool(no_rain_required),
        'estimated_cost': float(estimated_cost) if estimated_cost else None,
        'created_by': user
    }
    
    return event_data
